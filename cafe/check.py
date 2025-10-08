# -*- coding: utf-8 -*-
"""
네이버 카페 목록 중복 확인 도구
- store_name과 store_url_naver 기준으로 중복 체크
"""

from openpyxl import load_workbook
from collections import defaultdict

FILE_PATH = r"C:\Users\changjin\workspace\lab\pln\cafe\naver_cafe_list.xlsx"

def check_duplicates():
    print("🔍 중복 확인 시작...\n")
    
    try:
        wb = load_workbook(FILE_PATH)
        ws = wb.active
        
        # 데이터 읽기 (헤더 제외)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        total_count = len(rows)
        
        print(f"📊 총 {total_count}개 항목 발견\n")
        
        # 중복 체크용 딕셔너리
        name_dict = defaultdict(list)  # store_name -> [row_numbers]
        url_dict = defaultdict(list)   # store_url_naver -> [row_numbers]
        
        # 데이터 수집
        for idx, row in enumerate(rows, start=2):  # 엑셀 행 번호는 2부터 시작
            no, store_name, store_url_naver = row
            
            if store_name:
                name_dict[store_name].append((idx, no, store_url_naver))
            
            if store_url_naver:
                url_dict[store_url_naver].append((idx, no, store_name))
        
        # 중복 분석
        name_duplicates = {k: v for k, v in name_dict.items() if len(v) > 1}
        url_duplicates = {k: v for k, v in url_dict.items() if len(v) > 1}
        
        # 결과 출력
        print("="*70)
        print("📝 store_name 기준 중복")
        print("="*70)
        
        if name_duplicates:
            print(f"⚠️ {len(name_duplicates)}개의 중복된 이름 발견!\n")
            for name, occurrences in sorted(name_duplicates.items()):
                print(f"🔴 이름: {name}")
                print(f"   중복 횟수: {len(occurrences)}회")
                for row_num, no, url in occurrences:
                    print(f"   - 행 {row_num} (no={no}): {url}")
                print()
        else:
            print("✅ store_name 중복 없음!\n")
        
        print("="*70)
        print("🔗 store_url_naver 기준 중복")
        print("="*70)
        
        if url_duplicates:
            print(f"⚠️ {len(url_duplicates)}개의 중복된 URL 발견!\n")
            for url, occurrences in sorted(url_duplicates.items()):
                print(f"🔴 URL: {url}")
                print(f"   중복 횟수: {len(occurrences)}회")
                for row_num, no, name in occurrences:
                    print(f"   - 행 {row_num} (no={no}): {name}")
                print()
        else:
            print("✅ store_url_naver 중복 없음!\n")
        
        # 요약
        print("="*70)
        print("📈 요약")
        print("="*70)
        print(f"총 항목 수: {total_count}개")
        print(f"중복된 이름: {len(name_duplicates)}개")
        print(f"중복된 URL: {len(url_duplicates)}개")
        
        # 중복 제거 시 남을 항목 수
        unique_urls = len(url_dict)
        print(f"URL 기준 고유 항목: {unique_urls}개")
        print(f"중복 제거 시 삭제될 항목: {total_count - unique_urls}개")
        print("="*70)
        
    except FileNotFoundError:
        print(f"❌ 파일을 찾을 수 없습니다: {FILE_PATH}")
    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    check_duplicates()