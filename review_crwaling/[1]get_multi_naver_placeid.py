# -*- coding: utf-8 -*-
"""
네이버 지도 placeId 크롤러 (목록 전체 + '강릉' 주소 필터 + 행단위 저장)

- 검색 결과가 단일 라우팅이면 entryIframe에서 주소를 찾아 '강릉' 포함 시에만 저장
- 목록 화면이면 li 전부 훑어 주소 span 텍스트에 '강릉' 포함된 항목만 클릭 → entryIframe URL에서 placeId 추출
- 저장 스키마: [no, store_name, store_url_naver] 를 한 줄씩 append
"""

import os, re, time, urllib.parse
import pandas as pd
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException

# ===== 사용자 설정 =====
INPUT_PATH   = "강릉시_일반음식점 현황.csv"  # 입력 CSV
OUTPUT_PATH  = "naver_placeid_list.xlsx"     # 결과 엑셀
NAME_COL     = "업소명"                        # CSV에서 상호명 컬럼명
CITY_FILTER  = "강릉"                          # 주소에 이 문자열이 포함되어야 저장

# ===== 공통 유틸 =====
def read_korean_csv(path: str) -> pd.DataFrame:
    for enc in ("utf-8-sig", "cp949", "ms949", "euc-kr"):
        try:
            return pd.read_csv(path, encoding=enc)
        except UnicodeDecodeError:
            continue
    return pd.read_csv(path, encoding="utf-8", errors="ignore")

def build_review_url(place_id: str) -> str:
    # m.place 방문자 리뷰 URL
    return f"https://m.place.naver.com/restaurant/{place_id}/review/visitor?entry=ple&reviewSort=recent"

def extract_place_id_from_url(url: str) -> str | None:
    m = re.search(r"/place/(\d+)", url)
    if m:
        return m.group(1)
    m = re.search(r"[?&#]id=(\d+)", url)
    return m.group(1) if m else None

# ===== 저장: 행 단위 append (헤더 유지) =====
def append_single_row(path: str, no: int, store_name: str, store_url_naver: str):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["no", "store_name", "store_url_naver"])
        wb.save(path)
        print(f"📄 새 파일 생성 + 헤더 기록: {path}")

    wb = load_workbook(path)
    ws = wb.active
    ws.append([no, store_name, store_url_naver])
    wb.save(path)
    print(f"📝 저장완료 | no={no}, store_name='{store_name}'")

# ===== 프레임 헬퍼 =====
def _switch_to_search_iframe(driver, wait):
    driver.switch_to.default_content()
    wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, "searchIframe")))

def _switch_to_entry_iframe(driver, wait):
    driver.switch_to.default_content()
    wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, "entryIframe")))

# ===== 목록 스크롤/셀렉터 =====
def _scroll_all_in_list(driver):
    # 무한스크롤 컨테이너 후보
    candidates = [
        'div#_pcmap_list_scroll_container',
        'div#_pcmap_list_scroll_container div.api_scroller',
        'div[id$="_pcmap_list_scroll_container"]',
        'div#_pcmap_list_scroll_container ul'
    ]
    container = None
    for sel in candidates:
        try:
            container = driver.find_element(By.CSS_SELECTOR, sel)
            break
        except:
            pass
    if container is None:
        return
    prev_h = 0; stall = 0
    while True:
        driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight;", container)
        time.sleep(0.5)
        h = driver.execute_script("return arguments[0].scrollHeight;", container)
        stall = stall + 1 if h == prev_h else 0
        if stall >= 3:
            break
        prev_h = h

def _find_list_items(driver):
    # li 선택 (변동 대비 다중 셀렉터)
    li_selectors = [
        'div#_pcmap_list_scroll_container ul > li',
        'ul > li.UEzoS',
        'ul > li'
    ]
    for sel in li_selectors:
        lis = driver.find_elements(By.CSS_SELECTOR, sel)
        if lis:
            return lis
    return []

def _find_title_anchor(li):
    a_selectors = [
        'a.place_bluelink',
        'a.tit_name',
        'a[href*="/place/"]',
        'a[role="button"]',
        'a'
    ]
    for sel in a_selectors:
        try:
            return li.find_element(By.CSS_SELECTOR, sel)
        except:
            continue
    return None

# ===== 주소 추출 =====
def _extract_address_text_from_li(li) -> str:
    """
    목록 li 내부의 '주소' 텍스트를 찾아 반환.
    - 스샷 기준 후보: span.Pb4bU
    - 클래스 수시 개편 대비 폴백 포함
    """
    # 후보 우선
    for sel in ['span.Pb4bU', 'span[class*="addr"]', 'div[class*="address"] span']:
        spans = li.find_elements(By.CSS_SELECTOR, sel)
        for sp in spans:
            t = sp.text.strip()
            if t:
                return t
    # CITY_FILTER가 들어간 span 찾기
    for sp in li.find_elements(By.CSS_SELECTOR, "span"):
        t = sp.text.strip()
        if t and CITY_FILTER in t:
            return t
    # 시/군/구 포함 그럴싸한 텍스트
    for sp in li.find_elements(By.CSS_SELECTOR, "span"):
        t = sp.text.strip()
        if any(token in t for token in ("시", "군", "구")) and len(t) >= 4:
            return t
    return ""

def _extract_address_text_from_entry(driver) -> str:
    """
    상세(entryIframe)에서 주소 텍스트 추출(단일 라우팅용).
    구조가 종종 바뀌므로 넓게 탐색 후 CITY_FILTER 포함 텍스트를 우선 반환.
    """
    # 주소 영역에 흔한 패턴들
    candidates = [
        'span[class*="addr"]',
        'div[class*="address"]',
        'a[href^="https://map.naver.com/"] span',
        'div:has(> span) span'
    ]
    # 1) 후보 셀렉터 우선
    for sel in candidates:
        try:
            elems = driver.find_elements(By.CSS_SELECTOR, sel)
            texts = [e.text.strip() for e in elems if e.text.strip()]
            for t in texts:
                if CITY_FILTER in t:
                    return t
        except:
            pass
    # 2) 전체 span 중에서 CITY_FILTER 포함
    try:
        for sp in driver.find_elements(By.CSS_SELECTOR, "span"):
            t = sp.text.strip()
            if t and CITY_FILTER in t:
                return t
    except:
        pass
    return ""

# ===== 핵심: 검색 + 필터 + 수집 =====
def search_and_collect_filtered(driver, keyword: str, wait: WebDriverWait, settle_sec: float = 1.0):
    """
    반환: list[tuple(display_name, place_id, address)]
    """
    encoded = urllib.parse.quote(keyword, safe="")
    url = f"https://map.naver.com/p/search/{encoded}"
    driver.get(url)
    time.sleep(settle_sec)

    # 1) 단일 라우팅 감지
    try:
        WebDriverWait(driver, 5).until(lambda d: re.search(r"/place/\d+", d.current_url) is not None)
    except Exception:
        pass

    pid = extract_place_id_from_url(driver.current_url)
    if pid:
        # 단일 상세 → entryIframe 진입해서 주소 확인
        try:
            _switch_to_entry_iframe(driver, wait)
            addr = _extract_address_text_from_entry(driver)
        except TimeoutException:
            addr = ""
        if CITY_FILTER in addr:
            print(f"✅ 단일결과 저장 • {keyword} [{addr}] → {pid}")
            return [(keyword, pid, addr)]
        else:
            print(f"⛔ 단일결과 스킵(주소 불일치) • {keyword} [{addr}]")
            return []

    # 2) 목록 처리
    results = []
    try:
        _switch_to_search_iframe(driver, wait)
    except TimeoutException:
        print("⚠️ searchIframe 진입 실패")
        return results

    _scroll_all_in_list(driver)
    lis = _find_list_items(driver)
    if not lis:
        print("⚠️ 목록(li) 탐색 실패")
        return results

    print(f"📜 후보 {len(lis)}개 발견 (필터: '{CITY_FILTER}')")

    for idx in range(len(lis)):
        try:
            # stale 방지: 매회 재조회
            lis = _find_list_items(driver)
            li = lis[idx]

            addr = _extract_address_text_from_li(li)
            if CITY_FILTER not in addr:
                print(f"  • #{idx+1} 주소 미일치 → 스킵 | addr='{addr}'")
                continue

            # 표시 이름(가능하면 li 첫 줄)
            try:
                display_name = li.text.split("\n", 1)[0].strip()
            except:
                display_name = keyword

            a = _find_title_anchor(li)
            if a is None:
                print(f"  • #{idx+1} 링크 미발견 → 스킵")
                continue

            # 클릭하여 상세 진입 → entryIframe URL에서 id 추출
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", li)
            time.sleep(0.2)
            a.click()

            _switch_to_entry_iframe(driver, wait)
            WebDriverWait(driver, 8).until(lambda d: re.search(r"/place/\d+", d.current_url) is not None)
            current = driver.current_url
            pid = extract_place_id_from_url(current)
            if not pid:
                print(f"  • #{idx+1} {display_name} [{addr}] → placeId 파싱 실패")
                _switch_to_search_iframe(driver, wait)
                continue

            results.append((display_name, pid, addr))
            print(f"  ✅ 저장대상 • #{idx+1} {display_name} [{addr}] → {pid}")

            _switch_to_search_iframe(driver, wait)
            time.sleep(0.2)

        except (TimeoutException, StaleElementReferenceException) as e:
            try:
                _switch_to_search_iframe(driver, wait)
            except:
                pass
            print(f"  • #{idx+1} 오류: {e} → 계속")
            continue
        except Exception as e:
            print(f"  • #{idx+1} 알 수 없는 오류: {e} → 계속")
            try:
                _switch_to_search_iframe(driver, wait)
            except:
                pass
            continue

    return results

# ===== 메인 =====
def main():
    print("🚀 시작")
    df = read_korean_csv(INPUT_PATH)
    print(f"📋 대상 {len(df)}건")

    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    driver = webdriver.Chrome(options=options)
    wait = WebDriverWait(driver, 15)

    no = 1
    saved = skipped = 0

    for idx, row in df.iterrows():
        keyword = str(row.get(NAME_COL, "")).strip()
        if not keyword:
            continue

        print(f"\n{'='*60}\n🔎 [{idx+1}/{len(df)}] 검색: {keyword}")
        items = search_and_collect_filtered(driver, keyword, wait, settle_sec=2.0)

        if not items:
            print(f"❌ 저장 없음: {keyword}")
            skipped += 1
            continue

        for display_name, pid, addr in items:
            review_url = build_review_url(pid)
            append_single_row(OUTPUT_PATH, no, display_name, review_url)
            print(f"🎯 저장 • no={no} | {display_name} [{addr}] → {pid} | {review_url}")
            no += 1
            saved += 1
            time.sleep(0.2)

    driver.quit()
    print(f"\n🎉 완료: 저장 {saved}건, 스킵 {skipped}건")
    print(f"📄 결과 파일: {OUTPUT_PATH}")

if __name__ == "__main__":
    main()
