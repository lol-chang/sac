import re, time, os, urllib.parse
import pandas as pd
from openpyxl import load_workbook  # 엑셀 append 백업 용도 [설치 필요: pip install openpyxl]
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# ===== 경로 및 설정 =====
INPUT_PATH  = "강릉시_일반음식점 현황.csv"
OUTPUT_PATH = "naver_placeid_list"
NAME_COL    = "업소명"

# ===== 유틸 =====
def read_korean_csv(path: str) -> pd.DataFrame:
    for enc in ("utf-8-sig", "cp949", "ms949", "euc-kr"):
        try:
            return pd.read_csv(path, encoding=enc)
        except UnicodeDecodeError:
            continue
    return pd.read_csv(path, encoding="utf-8", errors="ignore")  # [9]

def build_review_url(place_id: str) -> str:
    # m.place 방문자 리뷰 URL
    return f"https://m.place.naver.com/restaurant/{place_id}/review/visitor?entry=ple&reviewSort=recent"

def extract_place_id_from_url(url: str) -> str | None:
    # 정규식은 raw-string으로, 단일 백슬래시 사용
    m = re.search(r"/place/(\d+)", url)  # [13][17]
    return m.group(1) if m else None

def append_to_excel(path: str, row_dict: dict):
    # 간단 병합 저장(파일 없으면 생성)
    df_new = pd.DataFrame([row_dict])
    if not os.path.exists(path):
        df_new.to_excel(path, index=False)
        print(f"📊 새 파일 생성: {path}")
        return
    try:
        existing_df = pd.read_excel(path)
        combined_df = pd.concat([existing_df, df_new], ignore_index=True)
        combined_df.to_excel(path, index=False)
        print("📊 엑셀 추가 완료")
    except Exception as e:
        # 백업 경로: openpyxl 직접 append
        try:
            book = load_workbook(path)
            with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                startrow = writer.sheets["Sheet1"].max_row
                df_new.to_excel(writer, index=False, header=False, startrow=startrow)
            print("📊 엑셀 append(백업) 완료")
        except Exception as ee:
            print(f"❌ 엑셀 저장 실패: {e} / 백업도 실패: {ee}")

# ===== 핵심: 검색 URL 직접 접근하여 place_id 추출 =====
def search_via_search_url_and_get_place_id(driver, name: str, wait: WebDriverWait, settle_sec: float = 1.0) -> str | None:
    """
    https://map.naver.com/p/search/{업체명} 로 직접 접근.
    단일 결과로 확정되면 현재 URL에 /place/{id}가 나타남 → place_id 반환.
    목록 화면이면 첫 결과 클릭 폴백 시도.
    """
    encoded = urllib.parse.quote(name, safe="")  # 한글/공백 안전 인코딩 [9][12]
    url = f"https://map.naver.com/p/search/{encoded}"
    driver.get(url)
    time.sleep(settle_sec)

    # 단일 결과 라우팅 감지: /place/{숫자}가 current_url에 등장하는지 확인
    try:
        WebDriverWait(driver, 5).until(lambda d: re.search(r"/place/\d+", d.current_url) is not None)
    except Exception:
        pass

    current = driver.current_url
    pid = extract_place_id_from_url(current)
    if pid:
        print(f"✅ 단일 결과 확정 → place_id: {pid} (from {current})")
        return pid

    # 폴백: 목록 화면일 수 있으므로 첫 결과 클릭 시도
    try:
        anchors = driver.find_elements(By.CSS_SELECTOR, "a.place_bluelink")
        if not anchors:
            anchors = driver.find_elements(By.CSS_SELECTOR, "a.a_item_click, a.UEzoS")
        if anchors:
            anchors.click()  # 리스트 중 첫 결과 클릭 [1][2]
            WebDriverWait(driver, 8).until(lambda d: re.search(r"/place/\d+", d.current_url))
            current = driver.current_url
            pid = extract_place_id_from_url(current)
            if pid:
                print(f"✅ 폴백 성공 → place_id: {pid} (from {current})")
                return pid
    except Exception as e:
        print(f"⚠️ 폴백 단계 오류: {e}")

    print("❌ place_id 추출 실패(검색 URL 경로)")
    return None

# ===== 메인 =====
def main():
    print("🚀 시작")
    df = read_korean_csv(INPUT_PATH)
    print(f"📋 대상 {len(df)}건")

    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    driver = webdriver.Chrome(options=options)
    wait = WebDriverWait(driver, 15)

    success = 0
    fail = 0
    no = 1

    for idx, row in df.iterrows():
        name = str(row.get(NAME_COL, "")).strip()
        if not name:
            continue

        print(f"\n{'='*50}\n🔎 [{idx+1}/{len(df)}] {name}")

        pid = search_via_search_url_and_get_place_id(driver, name, wait, settle_sec=2.0)
        if not pid:
            print(f"❌ {name} → place_id 실패")
            fail += 1
            time.sleep(0.5)
            continue

        review_url = build_review_url(pid)
        append_to_excel(OUTPUT_PATH, {
            "no": no,
            "store_name": name,
            "store_url_naver": review_url
        })
        print(f"✅ 저장: {review_url}")
        success += 1
        no += 1
        time.sleep(0.8)

    driver.quit()
    print(f"\n🎉 완료: 성공 {success}, 실패 {fail}")
    print(f"📄 결과 파일: {OUTPUT_PATH}")

if __name__ == "__main__":
    main()
